from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages 
from django.core.paginator import Paginator
import json
import datetime
from django.http import JsonResponse, HttpResponse
from userpreferences.models import Userpreferences
from .models import Source, Userincome

import csv
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# Create your views here.

def search_income(request):
    if request.method == 'POST':
        search_str = json.loads(request.body).get('searchText')
        income = Userincome.objects.filter(
            amount__istartswith=search_str, owner=request.user) | Userincome.objects.filter(
            date__istartswith=search_str, owner=request.user) | Userincome.objects.filter(
            description__icontains=search_str, owner=request.user) | Userincome.objects.filter(
            source__icontains=search_str, owner=request.user)
        data = income.values()
        return JsonResponse(list(data), safe=False)
    
@login_required(login_url='/authentication/login')
def index(request):
    categories = Source.objects.all()
    income = Userincome.objects.filter(owner = request.user)
    paginator = Paginator(income, 2)
    page_number = request.GET.get('page')
    page_obj = Paginator.get_page(paginator, page_number)
    currency = Userpreferences.objects.get(user=request.user).currency
    context = {
        "income": income,
        'page_obj': page_obj,
        'currency': currency
    }

    return render(request, 'income/index.html', context)

@login_required(login_url='/authentication/login')
def add_income(request):
    sources = Source.objects.all()
    context = {
        'sources': sources,
        'values': request.POST
    }

    if request.method == 'GET':
        return render(request, 'income/add_income.html', context)

    if request.method == 'POST':
        amount = request.POST.get('amount')
        description = request.POST.get('description')
        date = request.POST.get('income_date')
        source = request.POST.get('source')

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'income/add_income.html', context)

        if not description:
            messages.error(request, 'Description is required')
            return render(request, 'income/add_income.html', context)
        
        Userincome.objects.create(
            amount=amount,
            description=description,
            source=source,
            date=date,           
            owner=request.user
        )

        messages.success(request, 'Record added successfully')
        return redirect('income')

def Income_edit(request,id):
    income = Userincome.objects.get(pk=id)
    sources = Source.objects.all()
    context = {
        'income':income,
        'values': income,
        'sources':sources
    }
    if request.method == 'GET':
        return render(request, 'income/income-edit.html',context)
    
    if request.method == 'POST':
        amount = request.POST.get('amount')

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'income/income-edit.html', context)
        
    
        description = request.POST.get('description')
        date = request.POST.get('income_date')
        source = request.POST.get('source')
        

        if not description:
            messages.error(request, 'description is required')
            return render(request, 'income/income-edit.html', context)
        income.description=description
        income.amount=amount
        income.source=source
        income.date=date

        income.save()

        messages.success(request, 'Record updated successfully')
        return redirect('income')
    
def Income_delete(request, id):
    income = Userincome.objects.get(pk = id)
    income.delete()
    messages.success(request, 'income deleted') 
    return redirect('income')   

def income_summary(request):
    todays_date = datetime.date.today()
    six_months_ago = todays_date - datetime.timedelta(days=30 * 6)

    user_income= Userincome.objects.filter(
        owner=request.user,
        date__gte=six_months_ago,
        date__lte=todays_date
    )

    finalrep = {}

    for income in user_income:
        Source = income.source
        finalrep[Source] = finalrep.get(Source, 0) + income.amount

    return JsonResponse({'income_data': finalrep})

def statistics(request):
    return render(request, 'income/statistics.html')

def pdf_export(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=income.pdf'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph("Income Report", styles['Title'])
    elements.append(title)

    # Table header
    data = [['Amount', 'Source', 'Description', 'Date']]

    user_income= Userincome.objects.filter(owner=request.user)

    for income in user_income:
        data.append([
            str(income.amount),
            income.source,
            income.description,
            
            income.date.strftime('%Y-%m-%d')
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),

        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ]))

    elements.append(table)
    doc.build(elements)

    return response   

def csv_export(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        'attachment; filename=Income'
        + datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        + '.csv'
    )

    writer = csv.writer(response)
    writer.writerow(['Amount', 'Source','Description', 'Date'])

    user_income = Userincome.objects.filter(owner=request.user)

    for income in user_income:
        writer.writerow([
            income.amount,
            income.source,
            income.description,
            
            income.date
        ])

    return response

def excel_export(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income"

    # Header row
    headers = ['Amount','Source', 'Description', 'Date']
    ws.append(headers)

    # Fetch user Income
    user_income = Userincome.objects.filter(owner=request.user)

    for income in user_income:
        ws.append([
            income.amount,
            income.source,
            income.description,
            
            income.date
        ])

    # Auto-adjust column width
    for col_num, column_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=income.xlsx'

    wb.save(response)
    return response