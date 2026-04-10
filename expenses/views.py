from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from .models import Category, Expense
from django.contrib import messages 
from django.core.paginator import Paginator
import json
from django.http import JsonResponse,HttpResponse
from userpreferences.models import Userpreferences
import datetime
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.db.models import Q
from django.template.defaultfilters import date as django_date_filter
# Create your views here.
def search_expense(request):
    if request.method == 'POST':
        search_str = json.loads(request.body).get('searchText')
        expenses = Expense.objects.filter(
            amount__istartswith = search_str, owner=request.user    
        ) | Expense.objects.filter(
            date__istartswith = search_str, owner=request.user    
        ) | Expense.objects.filter(
            description__icontains = search_str, owner=request.user    
        ) | Expense.objects.filter(
            category__icontains = search_str, owner=request.user    
        )

        data = expenses.values()
        return JsonResponse(list(data), safe=False)
    
@login_required(login_url='/authentication/login')
def index(request):
    categories = Category.objects.all()
    expenses = Expense.objects.filter(owner=request.user)

    paginator = Paginator(expenses, 2)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    user_pref, created = Userpreferences.objects.get_or_create(
        user=request.user
    )

    context = {
        "expenses": expenses,
        "page_obj": page_obj,
        "currency": user_pref.currency
    }

    return render(request, 'expenses/index.html', context)

@login_required(login_url='/authentication/login')
def add_expense(request):
    categories = Category.objects.all()
    context = {
        'categories': categories,
        'values': request.POST
    }

    if request.method == 'GET':
        return render(request, 'expenses/add_expense.html', context)

    if request.method == 'POST':
        amount = request.POST.get('amount')

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'expenses/add_expense.html', context)
        
    
        description = request.POST.get('description')
        date = request.POST.get('expense_date')
        category = request.POST.get('category')
        

        if not description:
            messages.error(request, 'description is required')
            return render(request, 'expenses/add_expense.html', context)

       
        Expense.objects.create(
            amount=amount,
            description=description,
            category=category,
            date=date,
            owner=request.user
        )

        messages.success(request, 'Expense added successfully')
        return redirect('expenses')

def Expense_edit(request,id):
    expense = Expense.objects.get(pk=id)
    categories = Category.objects.all()
    context = {
        'expense':expense,
        'values': expense,
        'categories':categories
    }
    if request.method == 'GET':
        return render(request, 'expenses/edit-expense.html',context)
    
    if request.method == 'POST':
        amount = request.POST.get('amount')

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'expenses/edit-expense.html', context)
        
    
        description = request.POST.get('description')
        date = request.POST.get('expense_date')
        category = request.POST.get('category')
        

        if not description:
            messages.error(request, 'description is required')
            return render(request, 'expenses/edit-expense.html', context)

        
        expense.owner=request.user
        expense.description=description
        expense.amount=amount
        expense.category=category
        expense.date=date

        expense.save()

        messages.success(request, 'Expense added successfully')
        return redirect('expenses')
    
def Expense_delete(request, id):
    expense = Expense.objects.get(pk = id)
    expense.delete()
    messages.success(request, 'Expense deleted') 
    return redirect('expenses')   

from django.http import JsonResponse
import datetime


def expense_category_summary(request):
    todays_date = datetime.date.today()
    six_months_ago = todays_date - datetime.timedelta(days=30 * 6)

    expenses = Expense.objects.filter(
        owner=request.user,
        date__gte=six_months_ago,
        date__lte=todays_date
    )

    finalrep = {}

    for expense in expenses:
        category = expense.category
        finalrep[category] = finalrep.get(category, 0) + expense.amount

    return JsonResponse({'expense_category_data': finalrep})


def stats_view(request):
    return render(request, 'expenses/stats.html')

def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        'attachment; filename=Expenses'
        + datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        + '.csv'
    )

    writer = csv.writer(response)
    writer.writerow(['Amount', 'Description', 'Category', 'Date'])

    expenses = Expense.objects.filter(owner=request.user)

    for expense in expenses:
        writer.writerow([
            expense.amount,
            expense.description,
            expense.category,
            expense.date
        ])

    return response

def export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Header row
    headers = ['Amount', 'Description', 'Category', 'Date']
    ws.append(headers)

    # Fetch user expenses
    expenses = Expense.objects.filter(owner=request.user)

    for expense in expenses:
        ws.append([
            expense.amount,
            expense.description,
            expense.category,
            expense.date
        ])

    # Auto-adjust column width
    for col_num, column_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=expenses.xlsx'

    wb.save(response)
    return response
    
def export_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=expenses.pdf'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph("Expenses Report", styles['Title'])
    elements.append(title)

    # Table header
    data = [['Amount', 'Description', 'Category', 'Date']]

    expenses = Expense.objects.filter(owner=request.user)

    for expense in expenses:
        data.append([
            str(expense.amount),
            expense.description,
            expense.category,
            expense.date.strftime('%Y-%m-%d')
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),

        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ]))

    elements.append(table)
    doc.build(elements)

    return response   